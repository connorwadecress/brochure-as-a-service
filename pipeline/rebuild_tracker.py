"""
rebuild_tracker.py — Rebuild lead-tracker.xlsx with the correct column structure.

Run this ONCE to:
  1. Read all existing data rows (even if misaligned from the old schema)
  2. Create a fresh lead-tracker.xlsx with proper headers and styling
  3. Write migrated data into the correct columns

Old schema (14 cols, data written to cols 1-14):
  A(1)  Business Name  B(2)  Category     C(3)  Province    D(4)  Suburb
  E(5)  Contact Person F(6)  Phone        G(7)  Email       H(8)  FB/Insta
  I(9)  Has Website?   J(10) Lead Source  K(11) Status      L(12) Date Found
  M(13) Date Contacted N(14) Notes

New schema (17 cols):
  A(1)  Interest        B(2)  Business Name  C(3)  Category    D(4)  Province
  E(5)  City/Suburb     F(6)  Phone/WhatsApp G(7)  Email       H(8)  FB/Instagram
  I(9)  Maps URL        J(10) Rating         K(11) Reviews     L(12) Has Website?
  M(13) Lead Source     N(14) Status         O(15) Date Found  P(16) Date Contacted
  Q(17) Notes

Usage:
    cd pipeline
    python rebuild_tracker.py
    python rebuild_tracker.py --dry-run   # preview without writing
"""

import argparse
import shutil
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SCRIPT_DIR   = Path(__file__).parent
OLD_PATH     = SCRIPT_DIR / "lead-tracker.xlsx"
BACKUP_PATH  = SCRIPT_DIR / "lead-tracker.backup.xlsx"

DATA_START_ROW = 5

# ── Column definitions ────────────────────────────────────────────────────────

HEADERS = [
    "Interest",          # A  1
    "Business Name",     # B  2
    "Category",          # C  3
    "Province",          # D  4
    "City / Suburb",     # E  5
    "Phone / WhatsApp",  # F  6
    "Email",             # G  7
    "Facebook / Insta",  # H  8
    "Maps URL",          # I  9
    "Rating",            # J 10
    "Reviews",           # K 11
    "Has Website?",      # L 12
    "Lead Source",       # M 13
    "Status",            # N 14
    "Date Found",        # O 15
    "Date Contacted",    # P 16
    "Notes",             # Q 17
]

# Column widths in characters
COL_WIDTHS = [12, 30, 22, 15, 18, 18, 22, 30, 45, 8, 9, 12, 14, 14, 13, 15, 45]

# ── Styles ────────────────────────────────────────────────────────────────────

_THIN_SIDE   = Side(style="thin", color="E5E7EB")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE,  bottom=_THIN_SIDE,
)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F2937")  # dark slate
_HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
_SUB_FILL    = PatternFill(fill_type="solid", fgColor="374151")
_SUB_FONT    = Font(name="Arial", size=9, color="D1D5DB")
_TITLE_FONT  = Font(name="Arial", size=16, bold=True, color="1F2937")
_BODY_FONT   = Font(name="Arial", size=10, color="374151")

# Status colour map for the Status column (col 14)
STATUS_FILLS = {
    "New Lead":    PatternFill(fill_type="solid", fgColor="DBEAFE"),  # blue-100
    "Contacted":   PatternFill(fill_type="solid", fgColor="FEF3C7"),  # amber-100
    "Interested":  PatternFill(fill_type="solid", fgColor="D1FAE5"),  # green-100
    "Not Interested": PatternFill(fill_type="solid", fgColor="FEE2E2"),  # red-100
    "Site Created": PatternFill(fill_type="solid", fgColor="EDE9FE"),  # violet-100
}


# ── Migration helper ──────────────────────────────────────────────────────────

def _read_old_rows(ws) -> list[dict]:
    """Read existing data rows from the old (misaligned) schema.

    Old code wrote to columns 1-14:
      col1=name, col2=category, col3=province, col4=suburb, col5=contact,
      col6=phone, col7=email, col8=social, col9=has_website, col10=source,
      col11=status, col12=date_found, col13=date_contacted, col14=notes
    """
    rows = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if not name or str(name).startswith("Example:"):
            continue

        rows.append({
            "name":           str(name).strip(),
            "category":       ws.cell(row=r, column=2).value or "",
            "province":       ws.cell(row=r, column=3).value or "",
            "suburb":         ws.cell(row=r, column=4).value or "",
            "phone":          ws.cell(row=r, column=6).value or "",
            "email":          ws.cell(row=r, column=7).value or "",
            "social_url":     ws.cell(row=r, column=8).value or "",
            "has_website":    ws.cell(row=r, column=9).value or "No",
            "lead_source":    ws.cell(row=r, column=10).value or "Google Maps",
            "status":         ws.cell(row=r, column=11).value or "New Lead",
            "date_found":     ws.cell(row=r, column=12).value or date.today().isoformat(),
            "date_contacted": ws.cell(row=r, column=13).value or "",
            "notes":          ws.cell(row=r, column=14).value or "",
            # Fields not present in old schema — left blank
            "maps_url":  "",
            "rating":    "",
            "reviews":   "",
            "interest":  "",
        })
    return rows


# ── Builder ───────────────────────────────────────────────────────────────────

def _build_headers(ws) -> None:
    """Write rows 1-4: title, blank, column headers, sub-headers."""
    # Row 1: title
    ws.row_dimensions[1].height = 36
    title_cell = ws.cell(row=1, column=1, value="Lead Tracker — Brochure as a Service")
    title_cell.font      = _TITLE_FONT
    title_cell.alignment = Alignment(vertical="center")

    # Row 2: generation timestamp
    ws.row_dimensions[2].height = 18
    ts_cell = ws.cell(row=2, column=1,
                      value=f"Rebuilt: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ts_cell.font = Font(name="Arial", size=9, color="9CA3AF")

    # Row 3: column headers
    ws.row_dimensions[3].height = 28
    for col_idx, header in enumerate(HEADERS, 1):
        cell            = ws.cell(row=3, column=col_idx, value=header)
        cell.font       = _HEADER_FONT
        cell.fill       = _HEADER_FILL
        cell.border     = _THIN_BORDER
        cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Row 4: sub-header (usage hints)
    sub_hints = [
        "★ / Y / note",   # Interest
        "",               # Business Name
        "",               # Category
        "",               # Province
        "",               # City / Suburb
        "SA format",      # Phone
        "manual",         # Email
        "if only social", # FB / Insta
        "google.com/maps",# Maps URL
        "out of 5",       # Rating
        "count",          # Reviews
        "No / Yes",       # Has Website?
        "",               # Lead Source
        "New Lead / …",   # Status
        "YYYY-MM-DD",     # Date Found
        "YYYY-MM-DD",     # Date Contacted
        "opening hours",  # Notes
    ]
    ws.row_dimensions[4].height = 18
    for col_idx, hint in enumerate(sub_hints, 1):
        cell            = ws.cell(row=4, column=col_idx, value=hint)
        cell.font       = _SUB_FONT
        cell.fill       = _SUB_FILL
        cell.border     = _THIN_BORDER
        cell.alignment  = Alignment(horizontal="center", vertical="center")


def _write_data_rows(ws, rows: list[dict]) -> None:
    for row_idx, data in enumerate(rows, DATA_START_ROW):
        ws.row_dimensions[row_idx].height = 20
        row_values = [
            data["interest"],       # A (1)  Interest
            data["name"],           # B (2)  Business Name
            data["category"],       # C (3)  Category
            data["province"],       # D (4)  Province
            data["suburb"],         # E (5)  City / Suburb
            data["phone"],          # F (6)  Phone / WhatsApp
            data["email"],          # G (7)  Email
            data["social_url"],     # H (8)  Facebook / Insta
            data["maps_url"],       # I (9)  Maps URL
            data["rating"],         # J (10) Rating
            data["reviews"],        # K (11) Reviews
            data["has_website"],    # L (12) Has Website?
            data["lead_source"],    # M (13) Lead Source
            data["status"],         # N (14) Status
            data["date_found"],     # O (15) Date Found
            data["date_contacted"], # P (16) Date Contacted
            data["notes"],          # Q (17) Notes
        ]

        for col_idx, value in enumerate(row_values, 1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = _BODY_FONT
            cell.border    = _THIN_BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(col_idx == 17),  # Notes column
            )

        # Colour-code the Status cell
        status_val = data.get("status", "")
        if status_val in STATUS_FILLS:
            ws.cell(row=row_idx, column=14).fill = STATUS_FILLS[status_val]


def _apply_column_widths(ws) -> None:
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = width


def _freeze_panes(ws) -> None:
    ws.freeze_panes = ws["B5"]   # freeze header rows + Interest column


# ── Entry point ───────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Rebuild lead-tracker.xlsx")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview row count without writing files")
    args = parser.parse_args()

    if not OLD_PATH.exists():
        print(f"No existing tracker found at {OLD_PATH}")
        print("Creating a fresh empty tracker…")
        old_rows = []
    else:
        print(f"Reading existing tracker: {OLD_PATH}")
        wb_old = load_workbook(OLD_PATH, data_only=True)
        ws_old = wb_old["Lead Tracker"]
        old_rows = _read_old_rows(ws_old)
        print(f"  Found {len(old_rows)} existing data rows to migrate")

    if args.dry_run:
        print("\n[DRY RUN] Would create new tracker with:")
        for i, row in enumerate(old_rows[:5], 1):
            print(f"  {i}. {row['name']} | {row['category']} | {row['province']}")
        if len(old_rows) > 5:
            print(f"  … and {len(old_rows) - 5} more rows")
        return

    # Backup existing file
    if OLD_PATH.exists():
        shutil.copy2(OLD_PATH, BACKUP_PATH)
        print(f"  Backup saved to: {BACKUP_PATH}")

    # Build new workbook
    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = "Lead Tracker"

    _build_headers(ws_new)
    _write_data_rows(ws_new, old_rows)
    _apply_column_widths(ws_new)
    _freeze_panes(ws_new)

    wb_new.save(OLD_PATH)
    print(f"\nNew tracker saved: {OLD_PATH}")
    print(f"  Columns: {len(HEADERS)}")
    print(f"  Data rows migrated: {len(old_rows)}")
    print(f"\nColumn structure:")
    for i, h in enumerate(HEADERS, 1):
        col_letter = chr(ord('A') + i - 1)
        print(f"  {col_letter:2} ({i:2}): {h}")


if __name__ == "__main__":
    main()
