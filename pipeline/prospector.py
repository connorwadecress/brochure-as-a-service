"""
Lead Prospector — Brochure as a Service
========================================
Searches Google Maps for small SA businesses without websites
and logs them into the lead tracker spreadsheet.

Usage:
    python prospector.py                  # Auto-picks next suburb + industry from rotation
    python prospector.py --suburb Sandton --industry plumber   # Manual override
    python prospector.py --list           # Show the full rotation schedule
    python prospector.py --stats          # Show current tracker stats

Requires:
    pip install requests openpyxl
    A Google Places API key (free tier: $200/month credit)
    See README.md for setup instructions.
"""

import argparse
import json
import os
import sys
import requests
from datetime import datetime, date
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment

# ============================================================
# PATHS — Update these if you move the project folder
# ============================================================
SCRIPT_DIR = Path(__file__).parent
CONFIG_PATH = SCRIPT_DIR / "config.json"
TRACKER_PATH = SCRIPT_DIR / "lead-tracker.xlsx"
STATE_PATH = SCRIPT_DIR / ".prospector-state.json"

# ============================================================
# SEARCH ROTATION
# ============================================================
SUBURBS = {
    "Gauteng": [
        "Sandton", "Randburg", "Roodepoort", "Fourways", "Midrand",
        "Centurion", "Edenvale", "Benoni", "Boksburg", "Kempton Park",
        "Pretoria East", "Hatfield", "Menlyn", "Bryanston", "Northcliff"
    ],
    "Western Cape": [
        "Claremont", "Bellville", "Parow", "Table View", "Milnerton",
        "Durbanville", "Wynberg", "Fish Hoek", "Strand", "Somerset West",
        "Stellenbosch", "Paarl", "Constantia", "Sea Point", "Gardens"
    ],
    "KwaZulu-Natal": [
        "Umhlanga", "Ballito", "Pinetown", "Hillcrest", "Durban North",
        "Amanzimtoti", "Westville", "Chatsworth", "Phoenix", "Tongaat",
        "Pietermaritzburg", "Kloof", "La Lucia", "Berea", "Morningside"
    ]
}

INDUSTRIES = {
    "Trades & Home Services": [
        "plumber", "electrician", "painter", "cleaning service",
        "pest control", "locksmith", "handyman", "tiler",
        "waterproofing", "aircon installation", "garage door repair"
    ],
    "Beauty & Wellness": [
        "hair salon", "barber", "nail salon", "beauty salon",
        "spa", "lash extensions", "brow bar", "waxing salon"
    ],
    "Food & Hospitality": [
        "restaurant", "takeaway", "catering", "bakery",
        "butchery", "coffee shop", "food truck", "pizza"
    ],
    "Professional Services": [
        "tutor", "driving school", "photographer", "accountant",
        "event planner", "dog groomer", "vet", "nursery school"
    ]
}

# ============================================================
# LOAD CONFIG
# ============================================================
def load_config():
    if not CONFIG_PATH.exists():
        print("\n[ERROR] config.json not found!")
        print(f"Expected at: {CONFIG_PATH}")
        print("Run this first:  copy config.example.json config.json")
        print("Then paste your Google Places API key into it.\n")
        sys.exit(1)

    with open(CONFIG_PATH) as f:
        config = json.load(f)

    if not config.get("google_places_api_key") or config["google_places_api_key"] == "YOUR_API_KEY_HERE":
        print("\n[ERROR] No API key set in config.json!")
        print("Get one from: https://console.cloud.google.com/apis/credentials")
        print("Enable the 'Places API (New)' in your Google Cloud project.\n")
        sys.exit(1)

    return config


# ============================================================
# ROTATION STATE — tracks which combo to search next
# ============================================================
def load_state():
    if STATE_PATH.exists():
        with open(STATE_PATH) as f:
            return json.load(f)
    return {"rotation_index": 0, "runs": 0, "total_leads_found": 0}


def save_state(state):
    with open(STATE_PATH, "w") as f:
        json.dump(state, f, indent=2)


def _build_combos():
    """Builds a deterministically-shuffled list of all search combos.
    Shuffling interleaves cities & industries; the fixed seed keeps
    the order stable across runs so rotation_index stays valid."""
    combos = []
    for province, suburbs in SUBURBS.items():
        for suburb in suburbs:
            for category, keywords in INDUSTRIES.items():
                for keyword in keywords:
                    combos.append((suburb, keyword, province, category))
    import random as _rng
    _rng.Random(42).shuffle(combos)
    return combos


def get_next_search(state):
    """Picks the next search combo based on the rotation index."""
    combos = _build_combos()
    idx = state["rotation_index"] % len(combos)
    combo = combos[idx]

    # Advance for next run
    state["rotation_index"] = idx + 1
    return combo


# ============================================================
# GOOGLE PLACES API — Search + Details
# ============================================================
def search_places(api_key, query, max_results=10):
    """Use Places API Text Search to find businesses."""
    url = "https://places.googleapis.com/v1/places:searchText"
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": api_key,
        "X-Goog-FieldMask": ",".join([
            "places.id",
            "places.displayName",
            "places.formattedAddress",
            "places.nationalPhoneNumber",
            "places.internationalPhoneNumber",
            "places.websiteUri",
            "places.googleMapsUri",
            "places.rating",
            "places.userRatingCount",
            "places.businessStatus",
            "places.types",
            "places.regularOpeningHours"
        ])
    }
    body = {
        "textQuery": query,
        "maxResultCount": max_results,
        "languageCode": "en"
    }

    try:
        resp = requests.post(url, json=body, headers=headers, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        return data.get("places", [])
    except requests.exceptions.HTTPError as e:
        if resp.status_code == 403:
            print("\n[ERROR] API key rejected. Check that:")
            print("  1. Your API key is correct in config.json")
            print("  2. 'Places API (New)' is enabled in Google Cloud Console")
            print("  3. Billing is enabled on the project (free tier still needs it)\n")
        elif resp.status_code == 429:
            print("\n[WARN] Rate limited by Google. Wait a minute and try again.\n")
        else:
            print(f"\n[ERROR] API returned {resp.status_code}: {e}\n")
        return []
    except requests.exceptions.RequestException as e:
        print(f"\n[ERROR] Network error: {e}\n")
        return []


# ============================================================
# FILTER — Only businesses WITHOUT a website
# ============================================================
def filter_no_website(places):
    leads = []
    skipped = 0

    for place in places:
        website = place.get("websiteUri", "")
        status = place.get("businessStatus", "")

        # Skip if they have a website
        if website:
            skipped += 1
            continue

        # Skip if permanently closed
        if status == "CLOSED_PERMANENTLY":
            continue

        name = place.get("displayName", {}).get("text", "Unknown")
        phone = place.get("nationalPhoneNumber", place.get("internationalPhoneNumber", ""))
        address = place.get("formattedAddress", "")
        rating = place.get("rating", "")
        reviews = place.get("userRatingCount", 0)
        maps_url = place.get("googleMapsUri", "")

        # Build opening hours string
        hours_str = ""
        hours = place.get("regularOpeningHours", {})
        if hours and hours.get("weekdayDescriptions"):
            hours_str = " | ".join(hours["weekdayDescriptions"][:3])  # First 3 days to keep it short

        leads.append({
            "name": name,
            "phone": phone,
            "address": address,
            "rating": rating,
            "reviews": reviews,
            "maps_url": maps_url,
            "hours": hours_str
        })

    return leads, skipped


# ============================================================
# DUPLICATE CHECK — Don't add businesses already in the tracker
# ============================================================
def get_existing_names(ws):
    """Read all business names already in column A."""
    existing = set()
    for row in range(5, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val:
            existing.add(val.strip().lower())
    return existing


# ============================================================
# WRITE TO SPREADSHEET
# ============================================================
def append_to_tracker(leads, province, suburb, category):
    """Open the tracker spreadsheet and append new leads."""
    if not TRACKER_PATH.exists():
        print(f"\n[ERROR] Tracker not found at: {TRACKER_PATH}")
        print("Make sure lead-tracker.xlsx is in the pipeline folder.\n")
        return 0

    wb = load_workbook(TRACKER_PATH)
    ws = wb["Lead Tracker"]

    existing = get_existing_names(ws)
    added = 0

    # Find the first empty row
    next_row = ws.max_row + 1
    for r in range(5, ws.max_row + 1):
        if not ws.cell(row=r, column=1).value:
            next_row = r
            break

    # Styling
    body_font = Font(name="Arial", size=10, color="374151")
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    for lead in leads:
        # Skip duplicates
        if lead["name"].strip().lower() in existing:
            continue

        # Build notes string
        notes_parts = []
        if lead["rating"]:
            notes_parts.append(f"{lead['rating']} stars")
        if lead["reviews"]:
            notes_parts.append(f"{lead['reviews']} reviews")
        if lead["hours"]:
            notes_parts.append(lead["hours"])
        if lead["maps_url"]:
            notes_parts.append(lead["maps_url"])
        notes = " | ".join(notes_parts)

        row_data = [
            lead["name"],           # A: Business Name
            category,               # B: Industry
            province,               # C: Province
            suburb,                 # D: City / Suburb
            "",                     # E: Contact Person
            lead["phone"],          # F: Phone / WhatsApp
            "",                     # G: Email
            "",                     # H: Facebook / Insta
            "No",                   # I: Has Website?
            "Google Maps",          # J: Lead Source
            "New Lead",             # K: Status
            date.today().isoformat(), # L: Date Found
            "",                     # M: Date Contacted
            notes                   # N: Notes
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=(col_idx == 14))

        existing.add(lead["name"].strip().lower())
        next_row += 1
        added += 1

    wb.save(TRACKER_PATH)
    return added


# ============================================================
# STATS — Show current tracker state
# ============================================================
def show_stats():
    if not TRACKER_PATH.exists():
        print("Tracker not found.")
        return

    wb = load_workbook(TRACKER_PATH, data_only=True)
    ws = wb["Lead Tracker"]

    total = 0
    by_status = {}
    by_industry = {}
    by_province = {}

    for row in range(5, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name or name.startswith("Example:"):
            continue

        total += 1
        industry = ws.cell(row=row, column=2).value or "Unknown"
        province = ws.cell(row=row, column=3).value or "Unknown"
        status = ws.cell(row=row, column=11).value or "Unknown"

        by_status[status] = by_status.get(status, 0) + 1
        by_industry[industry] = by_industry.get(industry, 0) + 1
        by_province[province] = by_province.get(province, 0) + 1

    state = load_state()

    print("\n" + "=" * 50)
    print("  LEAD TRACKER STATS")
    print("=" * 50)
    print(f"\n  Total leads:        {total}")
    print(f"  Prospector runs:    {state.get('runs', 0)}")
    print(f"\n  By Status:")
    for s, c in sorted(by_status.items(), key=lambda x: -x[1]):
        print(f"    {s:<20} {c}")
    print(f"\n  By Industry:")
    for s, c in sorted(by_industry.items(), key=lambda x: -x[1]):
        print(f"    {s:<28} {c}")
    print(f"\n  By Province:")
    for s, c in sorted(by_province.items(), key=lambda x: -x[1]):
        print(f"    {s:<20} {c}")
    print()


# ============================================================
# LIST — Show upcoming rotation
# ============================================================
def show_rotation():
    state = load_state()
    combos = _build_combos()
    idx = state["rotation_index"] % len(combos)
    print(f"\nTotal combos in rotation: {len(combos)}")
    print(f"Current position: {idx}")
    print(f"\nNext 10 searches:\n")
    for i in range(10):
        pos = (idx + i) % len(combos)
        suburb, keyword, province, category = combos[pos]
        print(f"  {i+1:>2}. \"{keyword}\" in {suburb} ({province})")
    print()


# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="Lead Prospector — Find SA businesses without websites")
    parser.add_argument("--suburb", help="Override suburb to search (e.g. Sandton)")
    parser.add_argument("--industry", help="Override industry keyword (e.g. plumber)")
    parser.add_argument("--province", help="Province for manual search (Gauteng/Western Cape/KwaZulu-Natal)", default="Gauteng")
    parser.add_argument("--category", help="Category for manual search", default="Trades & Home Services")
    parser.add_argument("--max", type=int, default=15, help="Max results to fetch (default: 15)")
    parser.add_argument("--list", action="store_true", help="Show the next 10 searches in the rotation")
    parser.add_argument("--stats", action="store_true", help="Show current tracker statistics")
    parser.add_argument("--dry-run", action="store_true", help="Search but don't write to spreadsheet")
    args = parser.parse_args()

    if args.list:
        show_rotation()
        return

    if args.stats:
        show_stats()
        return

    config = load_config()
    state = load_state()
    api_key = config["google_places_api_key"]

    # Determine what to search
    if args.suburb and args.industry:
        suburb = args.suburb
        keyword = args.industry
        province = args.province
        category = args.category
    else:
        suburb, keyword, province, category = get_next_search(state)

    query = f"{keyword} in {suburb}, South Africa"

    print("\n" + "=" * 50)
    print("  LEAD PROSPECTOR")
    print("=" * 50)
    print(f"\n  Searching:   \"{keyword}\" in {suburb}")
    print(f"  Province:    {province}")
    print(f"  Category:    {category}")
    print(f"  Max results: {args.max}")
    print(f"  Date:        {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"\n  Calling Google Places API...")

    # Search
    places = search_places(api_key, query, max_results=args.max)

    if not places:
        print("  No results returned. Check your API key or try a different search.")
        save_state(state)
        return

    print(f"  Found {len(places)} businesses total")

    # Filter
    leads, skipped = filter_no_website(places)
    print(f"  Skipped {skipped} (already have websites)")
    print(f"  Qualified leads (no website): {len(leads)}")

    if not leads:
        print("\n  No new leads this run. All businesses already have websites.")
        print("  Try a different suburb or industry.\n")
        state["runs"] = state.get("runs", 0) + 1
        save_state(state)
        return

    # Print leads
    print(f"\n  {'='*50}")
    print(f"  LEADS FOUND")
    print(f"  {'='*50}\n")

    for i, lead in enumerate(leads, 1):
        rating_str = f"{lead['rating']}★ ({lead['reviews']} reviews)" if lead['rating'] else "No rating"
        print(f"  {i}. {lead['name']}")
        print(f"     Phone:   {lead['phone'] or 'Not listed'}")
        print(f"     Area:    {lead['address']}")
        print(f"     Rating:  {rating_str}")
        print()

    # Write to tracker
    if args.dry_run:
        print("  [DRY RUN] — Not writing to spreadsheet.\n")
        added = 0
    else:
        added = append_to_tracker(leads, province, suburb, category)
        print(f"  Added {added} new leads to tracker ({len(leads) - added} duplicates skipped)")

    # Update state
    state["runs"] = state.get("runs", 0) + 1
    state["total_leads_found"] = state.get("total_leads_found", 0) + added
    save_state(state)

    # Summary
    print(f"\n  {'='*50}")
    print(f"  RUN SUMMARY")
    print(f"  {'='*50}")
    print(f"  Search:          \"{keyword}\" in {suburb}")
    print(f"  Businesses found: {len(places)}")
    print(f"  Have website:     {skipped}")
    print(f"  No website:       {len(leads)}")
    print(f"  Added to tracker: {added}")
    print(f"  Total runs:       {state['runs']}")
    print(f"  Total leads ever: {state['total_leads_found']}")
    print()


if __name__ == "__main__":
    main()
