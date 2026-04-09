"""
lib/stats.py — Reporting (Single Responsibility: display tracker statistics)

Owns reading the tracker spreadsheet and printing aggregated stats.
No API calls, no state writes, no filtering logic. Pure display.
"""

import logging
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

SCRIPT_DIR      = Path(__file__).parent.parent
TRACKER_PATH    = SCRIPT_DIR / "lead-tracker.xlsx"
DATA_START_ROW  = 5
EXAMPLE_PREFIX  = "Example:"


# ── Public API ───────────────────────────────────────────────────────────────

def show_stats(state: dict) -> None:
    """Aggregate and print lead tracker statistics."""
    if not TRACKER_PATH.exists():
        print("Tracker not found.")
        return

    wb = load_workbook(TRACKER_PATH, data_only=True)
    ws = wb["Lead Tracker"]

    total       = 0
    by_status:   dict = {}
    by_industry: dict = {}
    by_province: dict = {}

    for row in range(DATA_START_ROW, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name or str(name).startswith(EXAMPLE_PREFIX):
            continue

        total += 1
        industry = ws.cell(row=row, column=3).value  or "Unknown"   # C: Category
        province = ws.cell(row=row, column=4).value  or "Unknown"   # D: Province
        status   = ws.cell(row=row, column=14).value or "Unknown"   # N: Status

        by_status[status]     = by_status.get(status, 0)     + 1
        by_industry[industry] = by_industry.get(industry, 0) + 1
        by_province[province] = by_province.get(province, 0) + 1

    _print_stats(total, state, by_status, by_industry, by_province)


# ── Private helpers ──────────────────────────────────────────────────────────

def _print_stats(
    total: int,
    state: dict,
    by_status:   dict,
    by_industry: dict,
    by_province: dict,
) -> None:
    print("\n" + "=" * 50)
    print("  LEAD TRACKER STATS")
    print("=" * 50)
    print(f"\n  Total leads:        {total}")
    print(f"  Prospector runs:    {state.get('runs', 0)}")

    print(f"\n  By Status:")
    for s, c in sorted(by_status.items(),   key=lambda x: -x[1]):
        print(f"    {s:<20} {c}")

    print(f"\n  By Industry:")
    for s, c in sorted(by_industry.items(), key=lambda x: -x[1]):
        print(f"    {s:<28} {c}")

    print(f"\n  By Province:")
    for s, c in sorted(by_province.items(), key=lambda x: -x[1]):
        print(f"    {s:<20} {c}")

    print()
