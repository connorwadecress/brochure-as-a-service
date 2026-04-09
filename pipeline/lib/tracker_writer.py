"""
lib/tracker_writer.py — Excel Lead Tracker I/O (Single Responsibility: spreadsheet writes)

Column layout (17 columns):
  A  (1)  Interest         — manual flag/notes by user (star, Y, etc.)
  B  (2)  Business Name    — lead["name"]
  C  (3)  Category         — industry category from rotation
  D  (4)  Province         — province from rotation
  E  (5)  City / Suburb    — suburb from rotation
  F  (6)  Phone / WhatsApp — lead["phone"]
  G  (7)  Email            — blank (Google Places doesn't provide)
  H  (8)  Facebook / Insta — lead["social_url"] (if their only "site" is social media)
  I  (9)  Maps URL         — lead["maps_url"]
  J  (10) Rating           — lead["rating"]
  K  (11) Reviews          — lead["reviews"]
  L  (12) Has Website?     — "No"
  M  (13) Lead Source      — "Google Maps"
  N  (14) Status           — "New Lead"
  O  (15) Date Found       — today's date
  P  (16) Date Contacted   — blank (manual)
  Q  (17) Notes            — opening hours
"""

import logging
from datetime import date
from pathlib import Path
from typing import List, Set

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

logger = logging.getLogger(__name__)

# ── Constants ────────────────────────────────────────────────────────────────

SCRIPT_DIR     = Path(__file__).parent.parent
TRACKER_PATH   = SCRIPT_DIR / "lead-tracker.xlsx"
DATA_START_ROW = 5   # rows 1-4 are headers
NAME_COLUMN    = 2   # B: Business Name
NOTES_COLUMN   = 17  # Q: Notes

# Pre-build reusable style objects (module-level constants for efficiency)
_BODY_FONT   = Font(name="Arial", size=10, color="374151")
_THIN_SIDE   = Side(style="thin", color="E5E7EB")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE,  bottom=_THIN_SIDE,
)


# ── Public API ───────────────────────────────────────────────────────────────

def get_existing_names(ws) -> Set[str]:
    """Return normalised (lowercase, stripped) business names already in column B."""
    existing: Set[str] = set()
    for row in range(DATA_START_ROW, ws.max_row + 1):
        val = ws.cell(row=row, column=NAME_COLUMN).value
        if val:
            existing.add(str(val).strip().lower())
    return existing


def append_leads(leads: List[dict], province: str, suburb: str, category: str) -> int:
    """Append qualified, non-duplicate leads to lead-tracker.xlsx.

    Returns the number of rows actually added.
    """
    if not TRACKER_PATH.exists():
        logger.error("Tracker not found at: %s", TRACKER_PATH)
        return 0

    try:
        wb = load_workbook(TRACKER_PATH)
    except PermissionError:
        logger.error("Cannot open tracker — close the file in Excel first.")
        return 0

    ws       = wb["Lead Tracker"]
    existing = get_existing_names(ws)
    added    = 0
    next_row = _find_next_empty_row(ws)

    for lead in leads:
        if lead["name"].strip().lower() in existing:
            continue

        _write_lead_row(ws, next_row, lead, province, suburb, category)
        existing.add(lead["name"].strip().lower())
        next_row += 1
        added    += 1

    wb.save(TRACKER_PATH)
    return added


# ── Private helpers ──────────────────────────────────────────────────────────

def _find_next_empty_row(ws) -> int:
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if not ws.cell(row=r, column=NAME_COLUMN).value:
            return r
    return ws.max_row + 1


def _write_lead_row(ws, row: int, lead: dict, province: str, suburb: str, category: str) -> None:
    row_data = [
        "",                           # A (1):  Interest         — user fills manually
        lead["name"],                 # B (2):  Business Name
        category,                     # C (3):  Category
        province,                     # D (4):  Province
        suburb,                       # E (5):  City / Suburb
        lead["phone"] or "",          # F (6):  Phone / WhatsApp
        "",                           # G (7):  Email
        lead.get("social_url", ""),   # H (8):  Facebook / Instagram
        lead["maps_url"] or "",       # I (9):  Maps URL
        lead["rating"] or "",         # J (10): Rating
        lead["reviews"] or "",        # K (11): Reviews
        "No",                         # L (12): Has Website?
        "Google Maps",                # M (13): Lead Source
        "New Lead",                   # N (14): Status
        date.today().isoformat(),     # O (15): Date Found
        "",                           # P (16): Date Contacted
        lead["hours"] or "",          # Q (17): Notes (opening hours)
    ]

    for col_idx, value in enumerate(row_data, 1):
        cell           = ws.cell(row=row, column=col_idx, value=value)
        cell.font      = _BODY_FONT
        cell.border    = _THIN_BORDER
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=(col_idx == NOTES_COLUMN),
        )
